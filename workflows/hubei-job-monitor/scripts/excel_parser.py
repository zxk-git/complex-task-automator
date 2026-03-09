#!/usr/bin/env python3
"""
湖北招聘监控 — Excel 岗位表解析器
自动识别表头，提取岗位信息，按学历筛选硕士岗位。
"""
import io
import re
import sys
import tempfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import cfg, get_output_dir, setup_logger, save_json, load_json, now_iso, get_http_session

log = setup_logger("excel_parser")

# 表头关键词 → 标准列名映射（长关键词优先匹配）
HEADER_MAPPING = [
    # 优先匹配更具体的关键词
    ("招聘单位名称", "employer"),
    ("招聘单位代码", "employer_code"),
    ("主管部门名称", "parent_dept"),
    ("主管部门代码", "parent_dept_code"),
    ("岗位所需专业", "major"),
    ("招聘计划人数", "headcount"),
    ("岗位名称", "position"),
    ("岗位代码", "code"),
    ("岗位描述", "description"),
    ("岗位类别", "category"),
    ("岗位等级", "level"),
    ("招聘计划", "headcount"),
    ("招聘人数", "headcount"),
    ("计划人数", "headcount"),
    ("单位名称", "employer"),
    ("招聘单位", "employer"),
    ("用人部门", "department"),
    ("专业工作经历", "experience"),
    ("其他条件", "other"),
    ("计划数", "headcount"),
    ("人数", "headcount"),
    ("学历", "education"),
    ("学位", "degree"),
    ("年龄", "age"),
    ("专业", "major"),
    ("备注", "remark"),
]


def find_header_row(ws) -> tuple[int, dict]:
    """
    扫描前 10 行，找到包含关键表头的行。
    返回 (header_row_number, {标准列名: 列索引})
    """
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_values = []
        for cell in ws[row_idx]:
            val = str(cell.value).strip() if cell.value else ""
            row_values.append(val)

        # 检查这行是否像表头（至少包含 2 个关键词）
        col_map = {}
        used_cols = set()  # 防止同一列被重复映射
        for col_idx, val in enumerate(row_values):
            if not val or col_idx in used_cols:
                continue
            for keyword, std_name in HEADER_MAPPING:
                if keyword in val and std_name not in col_map and col_idx not in used_cols:
                    col_map[std_name] = col_idx
                    used_cols.add(col_idx)
                    break

        if len(col_map) >= 2 and "education" in col_map:
            log.debug("找到表头行: %d, 映射 %d 列", row_idx, len(col_map))
            return row_idx, col_map

    return -1, {}


def parse_excel(content: bytes, source_title: str = "") -> list:
    """
    解析 Excel 岗位表，返回岗位列表。
    每个岗位: {position, employer, headcount, education, degree, major, age, ...}
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        log.error("无法打开 Excel: %s", e)
        return []

    all_jobs = []

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        header_row, col_map = find_header_row(ws)

        if header_row < 0:
            log.debug("Sheet '%s' 未找到表头，跳过", ws_name)
            continue

        log.debug("Sheet '%s': header=%d, cols=%s", ws_name, header_row, list(col_map.keys()))

        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            cells = [str(c.value).strip() if c.value else "" for c in row]

            # 跳过空行
            if not any(cells):
                continue

            job = {"_source_title": source_title}
            for std_name, col_idx in col_map.items():
                if col_idx < len(cells):
                    job[std_name] = cells[col_idx]
                else:
                    job[std_name] = ""

            # 如果 employer 是代码而 parent_dept 是名称，使用 parent_dept
            if job.get("employer_code") and not job.get("employer"):
                job["employer"] = job.get("employer_code", "")
            if job.get("employer", "").isdigit() and job.get("parent_dept"):
                # employer 是代码，用主管部门名称
                job["employer_display"] = job["parent_dept"]
            elif job.get("employer"):
                job["employer_display"] = job["employer"]

            # 跳过无效行（没有岗位名称或学历信息）
            if not job.get("position") and not job.get("education"):
                continue

            all_jobs.append(job)

    wb.close()
    return all_jobs


def filter_master_jobs(jobs: list) -> list:
    """筛选学历要求包含硕士/研究生的岗位（排除仅博士）"""
    edu_keywords = cfg("filter.education_keywords", ["硕士", "研究生"])
    result = []

    for job in jobs:
        edu = job.get("education", "")
        degree = job.get("degree", "")
        combined = f"{edu} {degree}"

        # 匹配硕士/研究生
        if any(kw in combined for kw in edu_keywords):
            # 如果是"博士研究生"且不含"硕士"，也算（因为user选了"硕士学历"）
            # 但如果只写"博士研究生"不含"硕士"，排除
            if "博士" in edu and "硕士" not in edu and "研究生" not in edu.replace("博士研究生", ""):
                continue
            result.append(job)

    return result


def download_and_parse_attachments(items: list) -> list:
    """下载各公告的 Excel 附件，解析岗位"""
    session = get_http_session()
    all_parsed_jobs = []

    for item in items:
        attachments = item.get("attachments", [])
        excel_atts = [a for a in attachments if a.get("type") in ("xlsx", "xls")]

        if not excel_atts:
            continue

        for att in excel_atts:
            url = att["url"]
            log.info("下载 Excel: %s", att.get("name", url)[:60])

            try:
                resp = session.get(url, timeout=cfg("http.timeout", 30))
                resp.raise_for_status()
            except Exception as e:
                log.error("下载失败: %s — %s", url[:60], e)
                continue

            jobs = parse_excel(resp.content, source_title=item.get("title", ""))

            # 为每个岗位附加来源信息
            for job in jobs:
                job["_announcement_title"] = item.get("title", "")
                job["_announcement_url"] = item.get("url", "")
                job["_announcement_date"] = item.get("date", "")
                job["_source_id"] = item.get("source_id", "")
                job["_attachment_name"] = att.get("name", "")

            all_parsed_jobs.extend(jobs)
            log.info("  解析到 %d 个岗位", len(jobs))

    return all_parsed_jobs


def format_job_summary(job: dict) -> str:
    """格式化单个岗位为可读文本"""
    parts = []
    if job.get("employer"):
        parts.append(f"📍 {job['employer']}")
    if job.get("position"):
        parts.append(f"岗位: {job['position']}")
    if job.get("headcount"):
        parts.append(f"人数: {job['headcount']}")
    if job.get("education"):
        parts.append(f"学历: {job['education']}")
    if job.get("degree"):
        parts.append(f"学位: {job['degree']}")
    if job.get("major"):
        major = job["major"]
        if len(major) > 50:
            major = major[:50] + "..."
        parts.append(f"专业: {major}")
    if job.get("age"):
        parts.append(f"年龄: {job['age']}")
    return " | ".join(parts)


def run():
    """
    独立运行：从 scrape-result.json 读取公告列表，下载附件，解析岗位，
    筛选硕士岗位，输出到 excel-jobs.json。
    """
    out_dir = get_output_dir()
    scrape_file = Path(out_dir) / "scrape-result.json"
    scrape_data = load_json(scrape_file, {"items": []})
    items = scrape_data.get("items", [])

    if not items:
        log.info("无公告数据")
        save_json(Path(out_dir) / "excel-jobs.json", {
            "timestamp": now_iso(),
            "total_jobs": 0,
            "master_jobs": 0,
            "items": [],
        })
        return

    # 下载并解析 Excel
    all_jobs = download_and_parse_attachments(items)
    log.info("共解析 %d 个岗位", len(all_jobs))

    # 筛选硕士岗位
    master_jobs = filter_master_jobs(all_jobs)
    log.info("硕士岗位: %d 个", len(master_jobs))

    result = {
        "timestamp": now_iso(),
        "total_jobs": len(all_jobs),
        "master_jobs": len(master_jobs),
        "items": master_jobs,
    }

    save_json(Path(out_dir) / "excel-jobs.json", result)


if __name__ == "__main__":
    run()
